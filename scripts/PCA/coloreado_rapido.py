# -*- coding: utf-8 -*-
"""
Created on Tue Apr  7 12:39:18 2026

@author: Usuario
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from sklearn.preprocessing import MinMaxScaler
import os
import numpy as np

#%%

# ====================== CONFIGURACIÓN ======================

'''carpeta donde está X_reconstruido.csv'''
ruta = r".................."

nombre_base = "X_reconstruido.csv"

#%%

excel_path = os.path.join(ruta, "X_reconstruido.xlsx")

# ====================== LEER CSV ======================
os.chdir(ruta)
df = pd.read_csv(nombre_base)

print(f"CSV cargado: {df.shape[0]} filas × {df.shape[1]} columnas")

# ====================== CREAR EXCEL ======================
with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    # Guardamos SIN encabezados (como en tu caso de "dibujo_gray")
    df.to_excel(writer, sheet_name='Sheet1', index=False, header=True)

# Cargar con openpyxl para aplicar formatos finos
wb = load_workbook(excel_path)
ws = wb.active

# ====================== ANCHO DE COLUMNAS ======================
ancho_columna = 0.08
for col in range(1, ws.max_column + 1):
    col_letter = get_column_letter(col)
    ws.column_dimensions[col_letter].width = ancho_columna

# ====================== ALTO DE FILAS ======================
alto_fila = 0.75
for row in range(1, ws.max_row + 1):
    ws.row_dimensions[row].height = alto_fila

# ====================== FORMATO CONDICIONAL ======================
color_scale_rule = ColorScaleRule(
    start_type='min',
    start_color='000000',   # Negro
    end_type='max',
    end_color='FFFFFF'      # Blanco
)

rango = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
ws.conditional_formatting.add(rango, color_scale_rule)

# ====================== GUARDAR ======================
wb.save(excel_path)

print("✅ Excel generado correctamente")
print(f"   Archivo: {excel_path}")
print(f"   Dimensiones: {df.shape[0]} filas × {df.shape[1]} columnas")